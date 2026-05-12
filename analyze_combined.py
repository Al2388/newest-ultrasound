"""
Combined acoustic + cycler + temperature analysis for a single experiment.

Reads three independently-recorded data streams from one cycling experiment:
  - acoustic A-scan HDF5      (this project's v2.0 schema)
  - Autolab cycler XLSX        (Time, WE(1).Potential, WE(1).Current, ...)
  - PicoLog temperature CSV   (HH:MM:SS, channel temperatures)

Aligns them on a common elapsed-time axis (hours from session start) and
plots a stacked multi-panel figure: ToF / Amplitude / Energy / Voltage+Current
/ Temperature.

Each external stream has an optional `offset_s` parameter so you can shift
its start to match the acoustic recording when the three loggers were not
started at exactly the same wall-clock moment.

Usage
-----
  python analyze_combined.py <h5_path> --cycler <xlsx> --temp <csv>
  python analyze_combined.py <h5_path> --cycler <xlsx> --temp <csv> --cycler-offset 120
"""
from __future__ import annotations

import argparse
import os
from pathlib import Path

import h5py
import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np


# =============================================================================
# Readers
# =============================================================================

def _hilbert_envelope(x: np.ndarray) -> np.ndarray:
    """Analytic-signal magnitude (Hilbert envelope) for a real 1-D array.

    Same algorithm as hs5_control.envelope_hilbert — reproduced locally to keep
    this analysis script independent of the acquisition module.
    """
    x = np.asarray(x, dtype=np.float32)
    n = x.size
    if n == 0:
        return x
    X = np.fft.fft(x)
    H = np.zeros(n, dtype=np.float32)
    if n % 2 == 0:
        H[0] = 1.0; H[1:n // 2] = 2.0; H[n // 2] = 1.0
    else:
        H[0] = 1.0; H[1:(n + 1) // 2] = 2.0
    return np.abs(np.fft.ifft(X * H)).astype(np.float32, copy=False)


def _envelope_peak_subsample(env: np.ndarray) -> float:
    """Return the envelope-peak position with parabolic sub-sample refinement.

    Standard quadratic-interpolation refinement: fit a parabola through the
    discrete-argmax bin and its two neighbours, return the fractional offset
    where the parabola's vertex lies. Resolution improves from ±0.5 samples
    (50 ns at fs=20 MHz) to roughly ±0.05 samples (5 ns) on smooth envelopes.
    """
    n = env.size
    if n < 3:
        return float(np.argmax(env))
    k = int(np.argmax(env))
    if k == 0 or k == n - 1:
        return float(k)
    y0, y1, y2 = float(env[k - 1]), float(env[k]), float(env[k + 1])
    denom = (y0 - 2.0 * y1 + y2)
    if abs(denom) < 1e-12:
        return float(k)
    delta = 0.5 * (y0 - y2) / denom
    # Clamp delta to its theoretical range to suppress occasional outliers
    delta = max(-1.0, min(1.0, delta))
    return float(k) + delta


def compute_envelope_tof_refined(h5_path: str) -> np.ndarray:
    """For every snapshot in an A-scan HDF5, compute the envelope-peak ToF
    with sub-sample parabolic refinement. Returns an array of floats in µs.

    DC offset is removed per snapshot (matches the worker's feature pipeline).
    """
    with h5py.File(h5_path, "r") as f:
        wf  = f["waveforms"][:]
        fs  = float(f.attrs.get("fs_hz", 20_000_000.0))
        gate_start = float(f.attrs.get("gate_us_start", 0.0))
    tof = np.empty(wf.shape[0], dtype=np.float32)
    for i in range(wf.shape[0]):
        v   = wf[i] - float(wf[i].mean())
        env = _hilbert_envelope(v)
        pk  = _envelope_peak_subsample(env)
        tof[i] = gate_start + pk / fs * 1e6
    return tof


def compute_tof_xcorr(
    h5_path: str,
    ref_n: int = 60,
    max_lag_samples: int = 40,
) -> np.ndarray:
    """Cross-correlation based ToF — the publication-standard method for battery
    acoustics. Uses the *whole envelope shape* against a reference, so argmax
    peak-hopping (when two competing echo peaks swap dominance) no longer
    produces step artifacts in the reported ToF.

    Algorithm
    ---------
    1. Build a reference envelope = mean of the first ref_n snapshots' envelopes
       (typically the first ~minute of recording, when coupling is fresh).
       The reference's own envelope-peak ToF anchors the absolute scale.
    2. For each snapshot:
         a. Compute its Hilbert envelope.
         b. Cross-correlate (zero-mean) with the reference envelope.
         c. Restrict the search to |lag| <= max_lag_samples so the xcorr
            cannot lock onto a far-off spurious correlation at the gate edges.
         d. Find the xcorr peak with parabolic sub-sample refinement.
         e. Report absolute ToF = reference ToF + lag.

    Parameters
    ----------
    ref_n            : how many opening snapshots to average into the reference
    max_lag_samples  : how far the lag is allowed to wander from zero in samples
                       (default 40 = 2 µs at fs=20 MHz, far more than any
                       physical SOC-driven shift)
    """
    with h5py.File(h5_path, "r") as f:
        wf = f["waveforms"][:]
        fs = float(f.attrs.get("fs_hz", 20_000_000.0))
        gate_start = float(f.attrs.get("gate_us_start", 0.0))

    n_rows, n_samp = wf.shape
    ref_n = max(1, min(ref_n, n_rows))

    # Build reference envelope by averaging the first ref_n envelopes
    ref_env_stack = np.empty((ref_n, n_samp), dtype=np.float32)
    for i in range(ref_n):
        v = wf[i] - float(wf[i].mean())
        ref_env_stack[i] = _hilbert_envelope(v)
    ref_env = ref_env_stack.mean(axis=0).astype(np.float32)
    ref_env_zm = ref_env - float(ref_env.mean())
    # Anchor: reference's own sub-sample envelope-peak ToF (absolute, gate-relative)
    ref_pk = _envelope_peak_subsample(ref_env)
    ref_tof_us = gate_start + ref_pk / fs * 1e6

    # Pre-compute the lag range we'll search through
    max_lag = max(1, min(int(max_lag_samples), n_samp // 2))
    center  = n_samp - 1                          # lag=0 location in 'full' xcorr
    lag_lo  = center - max_lag
    lag_hi  = center + max_lag + 1                # +1 for slice exclusivity

    tof = np.empty(n_rows, dtype=np.float32)
    for i in range(n_rows):
        v   = wf[i] - float(wf[i].mean())
        env = _hilbert_envelope(v)
        env_zm = env - float(env.mean())
        xc = np.correlate(env_zm, ref_env_zm, mode="full")
        window = xc[lag_lo:lag_hi]
        k_local = int(np.argmax(window))
        k_abs   = lag_lo + k_local

        # Parabolic refinement around the cross-correlation peak
        if 0 < k_abs < len(xc) - 1:
            y0, y1, y2 = float(xc[k_abs - 1]), float(xc[k_abs]), float(xc[k_abs + 1])
            denom = y0 - 2.0 * y1 + y2
            delta = 0.5 * (y0 - y2) / denom if abs(denom) > 1e-12 else 0.0
            delta = max(-1.0, min(1.0, delta))
        else:
            delta = 0.0

        lag_samples = (k_abs + delta) - center
        tof[i] = ref_tof_us + lag_samples / fs * 1e6

    return tof


def read_acoustic(h5_path: str, tof_method: str = "stored") -> dict:
    """Load per-snapshot acoustic features from a v2.0 A-scan HDF5 file.

    tof_method
    ----------
      "stored"   — use the integer-sample /tof_us already in the file (default)
      "envelope" — recompute envelope-peak ToF with parabolic sub-sample
                   refinement; smooths within-plateau noise but doesn't help
                   when two envelope peaks swap dominance (argmax hopping)
      "xcorr"    — cross-correlation ToF against a reference envelope built
                   from the first 60 snapshots; immune to argmax hopping,
                   reveals the smooth underlying SOC-driven shift
    """
    if tof_method not in ("stored", "envelope", "xcorr"):
        raise ValueError(f"Unknown tof_method: {tof_method!r}")

    with h5py.File(h5_path, "r") as f:
        ts = f["timestamps"][:]
        tof_name = "tof_us_absolute" if "tof_us_absolute" in f else "tof_us"
        tof = f[tof_name][:]
        if tof_name == "tof_us":
            tof = tof + float(f.attrs.get("gate_us_start", 0.0))
        out = {
            "t_s": ts - ts[0],
            "tof": tof,
            "amp": f["amplitude"][:],
            "eng": f["energy"][:],
        }
    if tof_method == "envelope":
        out["tof"] = compute_envelope_tof_refined(h5_path)
    elif tof_method == "xcorr":
        out["tof"] = compute_tof_xcorr(h5_path)
    return out


def read_cycler_xlsx(xlsx_path: str) -> dict:
    """Read an Autolab Nova export. Returns time, voltage, current arrays.

    Expects columns: Time (s) | WE(1).Potential (V) | ... | WE(1).Current (A) | ...
    Reads via openpyxl in non-readonly mode so the row count is reliable.
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header = [c.value for c in ws[1]]
    idx_t = header.index("Time (s)")
    idx_v = header.index("WE(1).Potential (V)")
    idx_i = header.index("WE(1).Current (A)")

    t, v, i = [], [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx_t] is None:
            continue
        try:
            t.append(float(row[idx_t]))
            v.append(float(row[idx_v]))
            i.append(float(row[idx_i]))
        except (TypeError, ValueError):
            continue

    t = np.asarray(t, dtype=np.float64)
    return {
        "t_s":  t - t[0] if t.size else t,   # relative to first row
        "v":    np.asarray(v, dtype=np.float32),
        "i":    np.asarray(i, dtype=np.float32),
    }


def read_temperature_csv(csv_path: str) -> dict:
    """Read a PicoLog export. Returns time and temperature arrays.

    Expected format: header row, then rows of `"HH:MM:SS","temperature"`.
    Trailing blank-temperature rows are skipped. Time is converted to seconds
    relative to the first valid row.
    """
    import csv
    t_s, temp = [], []
    with open(csv_path) as f:
        reader = csv.reader(f)
        next(reader, None)  # skip header
        for row in reader:
            if len(row) < 2 or row[1] == "" or row[0] == "":
                continue
            try:
                hh, mm, ss = row[0].split(":")
                seconds = int(hh) * 3600 + int(mm) * 60 + int(ss)
                temp_val = float(row[1])
            except (ValueError, AttributeError):
                continue
            t_s.append(seconds)
            temp.append(temp_val)

    t_s = np.asarray(t_s, dtype=np.float64)
    return {
        "t_s":  t_s - t_s[0] if t_s.size else t_s,
        "temp": np.asarray(temp, dtype=np.float32),
    }


# =============================================================================
# Helpers
# =============================================================================

def _rolling_mean(x: np.ndarray, window: int) -> np.ndarray:
    """Centred rolling mean. Window=1 returns x; edges become NaN."""
    if window <= 1:
        return x.astype(np.float32, copy=True)
    kernel = np.ones(window, dtype=np.float64) / window
    smoothed = np.convolve(x.astype(np.float64), kernel, mode="same")
    half = window // 2
    smoothed[:half] = np.nan
    smoothed[-half:] = np.nan
    return smoothed.astype(np.float32)


# =============================================================================
# Plotting
# =============================================================================

def plot_combined(
    h5_path: str,
    cycler_xlsx: str | None = None,
    temperature_csv: str | None = None,
    cycler_offset_s: float = 0.0,
    temp_offset_s: float = 0.0,
    out_path: str | None = None,
    smooth_window_s: float = 60.0,
    figsize: tuple[float, float] = (13, 12),
    dpi: int = 140,
    t_min_h: float | None = None,
    t_max_h: float | None = None,
    tof_method: str = "stored",
) -> str:
    """
    Generate a stacked time-aligned plot combining acoustic, cycler, and
    temperature data for one experiment. Returns the saved PNG path.

    Panels (top → bottom): ToF, Amplitude, Energy, Voltage+Current, Temperature.
    Empty panels are skipped if the corresponding source file is not provided.
    """
    ac = read_acoustic(h5_path, tof_method=tof_method)
    cy = read_cycler_xlsx(cycler_xlsx) if cycler_xlsx else None
    tp = read_temperature_csv(temperature_csv) if temperature_csv else None

    # Common axis: elapsed hours from acoustic session start
    t_ac_h = ac["t_s"] / 3600.0
    if cy is not None:
        t_cy_h = (cy["t_s"] + cycler_offset_s) / 3600.0
    if tp is not None:
        t_tp_h = (tp["t_s"] + temp_offset_s) / 3600.0

    # Smoothing for acoustic — temperature and cycler are usually already smooth enough
    dt_ac = float(np.median(np.diff(ac["t_s"]))) if len(ac["t_s"]) > 1 else 1.0
    win = max(1, int(round(smooth_window_s / max(dt_ac, 1e-6))))
    do_smooth = smooth_window_s > 0 and win > 1

    # Build panels list — skip the ones with no data
    panels = []
    tof_label = {
        "stored":   "ToF (us)",
        "envelope": "ToF (us, refined)",
        "xcorr":    "ToF (us, xcorr)",
    }[tof_method]
    panels.append((tof_label,        t_ac_h, ac["tof"], "tab:red"))
    panels.append(("Amplitude (V)",  t_ac_h, ac["amp"], "tab:green"))
    panels.append(("Energy",         t_ac_h, ac["eng"], "tab:purple"))
    if cy is not None:
        panels.append(("Cycler",     None,    cy,        None))
    if tp is not None:
        panels.append(("Temperature (°C)", t_tp_h, tp["temp"], "tab:orange"))

    n_panels = len(panels)
    fig, axes = plt.subplots(n_panels, 1, figsize=figsize, dpi=dpi, sharex=True)
    if n_panels == 1:
        axes = [axes]

    title_bits = [
        f"{Path(h5_path).name}",
        f"{len(ac['t_s']):,} acoustic snapshots, {t_ac_h[-1]:.2f} h",
    ]
    if cy is not None:
        title_bits.append(f"cycler {t_cy_h[-1]-t_cy_h[0]:.1f}h")
    if tp is not None:
        title_bits.append(f"temp {t_tp_h[-1]-t_tp_h[0]:.1f}h")
    if do_smooth:
        title_bits.append(f"smoothed {smooth_window_s:.0f}s")
    fig.suptitle("  -  ".join(title_bits), fontsize=11)

    for ax, panel in zip(axes, panels):
        name = panel[0]
        if name == "Cycler":
            # Two-axis panel: voltage on left, current on right
            cy_data = panel[2]
            t_cy = t_cy_h
            ax.plot(t_cy, cy_data["v"], color="tab:blue", linewidth=1.0, label="V")
            ax.set_ylabel("Voltage (V)", color="tab:blue")
            ax.tick_params(axis="y", labelcolor="tab:blue")

            ax2 = ax.twinx()
            ax2.plot(t_cy, cy_data["i"] * 1000.0, color="tab:gray",
                     linewidth=1.0, label="I (mA)", alpha=0.7)
            ax2.set_ylabel("Current (mA)", color="tab:gray")
            ax2.tick_params(axis="y", labelcolor="tab:gray")
            ax2.axhline(0.0, color="tab:gray", linewidth=0.4, alpha=0.3)
        else:
            label, t_arr, y_arr, color = panel
            label_root = label.split(" (")[0]
            if do_smooth and label_root in ("ToF", "Amplitude", "Energy"):
                ax.plot(t_arr, y_arr, color=color, linewidth=0.4, alpha=0.25, label="raw")
                ax.plot(t_arr, _rolling_mean(y_arr, win), color=color, linewidth=1.4, label="smoothed")
            else:
                ax.plot(t_arr, y_arr, color=color, linewidth=1.0)
            ax.set_ylabel(label)
        ax.grid(True, alpha=0.25)

    axes[-1].set_xlabel("elapsed time (hours from acoustic session start)")
    # Apply legend only to top acoustic panel
    if do_smooth:
        axes[0].legend(loc="upper right", fontsize=8)

    # Apply x-axis time window if requested. Y-limits are recomputed inside the
    # window so a trimmed plot isn't squished by out-of-window outliers.
    if t_min_h is not None or t_max_h is not None:
        lo = t_min_h if t_min_h is not None else min(t_ac_h.min(),
            t_tp_h.min() if tp is not None else float("inf"))
        hi = t_max_h if t_max_h is not None else t_ac_h.max()
        for ax, panel in zip(axes, panels):
            ax.set_xlim(lo, hi)
            # Recompute y-limits using only data within the visible window.
            # The Cycler panel is a special case (two-axis); handle separately.
            name = panel[0]
            if name == "Cycler":
                cy_data = panel[2]
                mask = (t_cy_h >= lo) & (t_cy_h <= hi)
                if mask.any():
                    vw = cy_data["v"][mask]
                    iw = cy_data["i"][mask] * 1000.0
                    ax.set_ylim(vw.min() - 0.02, vw.max() + 0.02)
                    # twin axis: find via shared get_shared_x_axes is messy;
                    # iterate fig.axes for the same x position
                    for twin in [a for a in fig.axes
                                 if a is not ax and a.bbox.bounds == ax.bbox.bounds]:
                        twin.set_ylim(iw.min() - 2, iw.max() + 2)
            else:
                _, t_arr, y_arr, _ = panel
                mask = (t_arr >= lo) & (t_arr <= hi) & np.isfinite(y_arr)
                if mask.any():
                    yw = y_arr[mask]
                    pad = (yw.max() - yw.min()) * 0.05 if yw.max() > yw.min() else 0.01
                    ax.set_ylim(yw.min() - pad, yw.max() + pad)

    fig.tight_layout(rect=(0, 0, 1, 0.97))

    if out_path is None:
        out_path = str(Path(h5_path).with_suffix("")) + "_combined.png"
    fig.savefig(out_path, bbox_inches="tight")
    plt.close(fig)
    return os.path.abspath(out_path)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("h5", help="A-scan HDF5 file")
    parser.add_argument("--cycler", default=None, help="Autolab Nova XLSX export")
    parser.add_argument("--temp",   default=None, help="PicoLog temperature CSV")
    parser.add_argument("--cycler-offset", type=float, default=0.0,
                        help="Shift cycler time by this many seconds")
    parser.add_argument("--temp-offset",   type=float, default=0.0,
                        help="Shift temperature time by this many seconds")
    parser.add_argument("--smooth", type=float, default=60.0,
                        help="Rolling-mean window (s) for acoustic traces; 0 disables")
    parser.add_argument("--t-min", type=float, default=None,
                        help="Lower x-axis bound in hours from acoustic start")
    parser.add_argument("--t-max", type=float, default=None,
                        help="Upper x-axis bound in hours from acoustic start")
    tof_group = parser.add_mutually_exclusive_group()
    tof_group.add_argument("--envelope-tof", action="store_true",
                           help="Recompute ToF from /waveforms with parabolic sub-sample "
                                "refinement (cures 50 ns quantization, but not peak-hopping).")
    tof_group.add_argument("--xcorr-tof", action="store_true",
                           help="Recompute ToF by cross-correlation against a reference "
                                "envelope built from the first 60 snapshots. "
                                "Immune to argmax peak-hopping when two echo peaks "
                                "swap dominance; reveals the smooth SOC-driven shift.")
    parser.add_argument("--out", default=None, help="Output PNG path")
    args = parser.parse_args()

    # Quick stats printout per source
    ac = read_acoustic(args.h5)
    print(f"acoustic:   {len(ac['t_s']):,} rows  duration={ac['t_s'][-1]/3600:.2f}h")
    if args.cycler:
        cy = read_cycler_xlsx(args.cycler)
        print(f"cycler:     {len(cy['t_s']):,} rows  duration={cy['t_s'][-1]/3600:.2f}h  "
              f"V range=[{cy['v'].min():.3f}, {cy['v'].max():.3f}] V  "
              f"I range=[{cy['i'].min()*1000:.1f}, {cy['i'].max()*1000:.1f}] mA")
    if args.temp:
        tp = read_temperature_csv(args.temp)
        print(f"temperature:{len(tp['t_s']):,} rows  duration={tp['t_s'][-1]/3600:.2f}h  "
              f"T range=[{tp['temp'].min():.2f}, {tp['temp'].max():.2f}] C")

    out = plot_combined(
        args.h5,
        cycler_xlsx=args.cycler,
        temperature_csv=args.temp,
        cycler_offset_s=args.cycler_offset,
        temp_offset_s=args.temp_offset,
        out_path=args.out,
        smooth_window_s=args.smooth,
        t_min_h=args.t_min,
        t_max_h=args.t_max,
        tof_method=("xcorr" if args.xcorr_tof
                    else "envelope" if args.envelope_tof
                    else "stored"),
    )
    print(f"\nsaved plot: {out}")


if __name__ == "__main__":
    main()

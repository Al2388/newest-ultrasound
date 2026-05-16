"""
End-to-end cycle analysis pipeline.

Synchronises three independently-logged streams on the cycler clock,
computes SoC from Coulomb counting, extracts NCC-based ToF and a
windowed attenuation feature from each averaged A-scan, and produces
the standard diagnostic plot set plus a quantitative noise summary.

Pipeline (matches user spec):
  1. Sync three streams onto cycler timestamps (linear interp for temp,
     nearest-snapshot match for A-scan).
  2. SoC(t) = SoC_0 + (1/Q_nominal) * integral(I dt) * 100.
  3. ToF per scan via normalised cross-correlation against a reference
     scan, refined to sub-sample precision with a parabolic fit.
     Attenuation per scan via peak-to-peak voltage of the back-wall echo
     window (also reports windowed energy).
  4. Diagnostic plots (saved as separate PNGs at project root):
       diag_voltage_current.png
       diag_temperature.png
       diag_tof_vs_time.png
       headline_tof_vs_soc.png
       attenuation_vs_soc.png
       cycle_drift.png
  5. Noise / sensitivity summary printed to stdout.

Usage:
  python cycle_analysis.py <h5_path> --cycler <xlsx> --temp <csv> \\
        [--q-nominal-ah 0.86] [--soc-initial-pct 0] [--ref-snapshot 0] \\
        [--echo-window-us 30 40] [--cycler-offset 0] [--temp-offset 0]
"""
from __future__ import annotations

import argparse
import os
from pathlib import Path

import h5py
import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np


# =============================================================================
# Readers (same conventions as analyze_combined.py)
# =============================================================================

def read_acoustic(h5_path: str) -> dict:
    with h5py.File(h5_path, "r") as f:
        ts = f["timestamps"][:]
        wf = f["waveforms"][:]
        fs = float(f.attrs.get("fs_hz", 20_000_000.0))
        gate_us_start = float(f.attrs.get("gate_us_start", 30.0))
        gate_us_end   = float(f.attrs.get("gate_us_end",   40.0))
    return {
        "t_unix": ts,
        "t_s":    ts - ts[0],
        "wf":     wf,
        "fs":     fs,
        "gate_us_start": gate_us_start,
        "gate_us_end":   gate_us_end,
    }


def read_cycler_xlsx(xlsx_path: str) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in ws[1]]
    idx_t = header.index("Time (s)")
    idx_v = header.index("WE(1).Potential (V)")
    idx_i = header.index("WE(1).Current (A)")
    t, v, i = [], [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx_t] is None:
            continue
        try:
            t.append(float(row[idx_t]))
            v.append(float(row[idx_v]))
            i.append(float(row[idx_i]))
        except (TypeError, ValueError):
            continue
    t = np.asarray(t, dtype=np.float64)
    return {
        "t_s": t - t[0],
        "v":   np.asarray(v, dtype=np.float32),
        "i":   np.asarray(i, dtype=np.float32),
    }


def read_temperature_csv(csv_path: str) -> dict:
    import csv
    t_s, temp = [], []
    with open(csv_path) as f:
        reader = csv.reader(f)
        next(reader, None)
        for row in reader:
            if len(row) < 2 or row[1] == "" or row[0] == "":
                continue
            try:
                hh, mm, ss = row[0].split(":")
                seconds = int(hh) * 3600 + int(mm) * 60 + int(ss)
                temp_val = float(row[1])
            except (ValueError, AttributeError):
                continue
            t_s.append(seconds)
            temp.append(temp_val)
    t_s = np.asarray(t_s, dtype=np.float64)
    return {
        "t_s":  t_s - t_s[0] if t_s.size else t_s,
        "temp": np.asarray(temp, dtype=np.float32),
    }


# =============================================================================
# Signal helpers
# =============================================================================

def _parabolic_peak(y: np.ndarray, k: int) -> float:
    """Sub-sample peak position via 3-point parabolic fit. Clamped to ±1."""
    if y.size < 3 or k <= 0 or k >= y.size - 1:
        return float(k)
    y0, y1, y2 = float(y[k - 1]), float(y[k]), float(y[k + 1])
    denom = y0 - 2.0 * y1 + y2
    if abs(denom) < 1e-12:
        return float(k)
    delta = 0.5 * (y0 - y2) / denom
    return float(k) + max(-1.0, min(1.0, delta))


def ncc_tof_shift(scan: np.ndarray, ref: np.ndarray, dt: float) -> tuple[float, float]:
    """Normalised cross-correlation ToF shift with parabolic refinement.

    Returns (shift_seconds, peak_correlation). Positive shift means `scan`
    is delayed relative to `ref` (echo arrives later).
    """
    s = scan - scan.mean()
    r = ref  - ref.mean()
    norm = float(np.sqrt(np.sum(s * s) * np.sum(r * r)))
    if norm <= 1e-12:
        return float("nan"), float("nan")
    nc = np.correlate(s, r, mode="full") / norm
    center = len(r) - 1
    k = int(np.argmax(nc))
    k_refined = _parabolic_peak(nc, k)
    return float((k_refined - center) * dt), float(nc[k])


# =============================================================================
# Pipeline
# =============================================================================

def synchronise(ac: dict, cy: dict, tp: dict,
                cycler_offset_s: float = 0.0,
                temp_offset_s:   float = 0.0) -> dict:
    """Resample temperature onto cycler timestamps; map each A-scan to its
    nearest cycler row. All offsets are interpreted relative to the A-scan
    session start (t=0)."""
    t_cy = cy["t_s"] + cycler_offset_s     # cycler time in A-scan-elapsed frame
    t_tp = tp["t_s"] + temp_offset_s
    t_ac = ac["t_s"]

    # Temperature resampled onto cycler timestamps (linear interp + edge clamp)
    temp_on_cy = np.interp(t_cy, t_tp, tp["temp"],
                           left=tp["temp"][0], right=tp["temp"][-1])

    # Map each cycler row to its nearest A-scan snapshot (NCC ToF needs scan_i,
    # so we keep the snapshot index per cycler row).
    ac_idx_per_cy = np.clip(np.searchsorted(t_ac, t_cy), 0, len(t_ac) - 1)
    # Refine: searchsorted gives insertion point; pick the closer of (k-1, k)
    left  = np.clip(ac_idx_per_cy - 1, 0, len(t_ac) - 1)
    right = ac_idx_per_cy
    use_left = np.abs(t_ac[left] - t_cy) < np.abs(t_ac[right] - t_cy)
    ac_idx_per_cy = np.where(use_left, left, right)

    return {
        "t_cycler_h":  t_cy / 3600.0,
        "v":           cy["v"],
        "i":           cy["i"],
        "temp_on_cy":  temp_on_cy,
        "ac_idx":      ac_idx_per_cy,
        "t_ac_per_cy_h": t_ac[ac_idx_per_cy] / 3600.0,
    }


def coulomb_soc(t_s: np.ndarray, i_a: np.ndarray,
                q_nominal_ah: float, soc_initial_pct: float = 0.0,
                charge_positive: bool = True) -> np.ndarray:
    """SoC(t) = SoC_0 + (1/Q_nominal) * integral(I dt) * 100.
    With charge_positive=True, positive current adds SoC."""
    sign = 1.0 if charge_positive else -1.0
    q_ah = np.zeros_like(i_a, dtype=np.float64)
    if len(i_a) > 1:
        # Trapezoidal integration of I dt; convert seconds → hours
        dt_h = np.diff(t_s) / 3600.0
        i_mid = 0.5 * (i_a[1:] + i_a[:-1]) * sign
        q_ah[1:] = np.cumsum(i_mid * dt_h)
    return soc_initial_pct + (q_ah / q_nominal_ah) * 100.0


def compute_tof_shift_per_cycler(ac: dict, ac_idx: np.ndarray,
                                  ref_snapshot: int = 0) -> np.ndarray:
    """Per cycler row, NCC ToF shift (in seconds) of the corresponding A-scan
    relative to the reference scan."""
    dt = 1.0 / ac["fs"]
    ref = ac["wf"][ref_snapshot] - ac["wf"][ref_snapshot].mean()
    # Pre-compute unique scan indices to avoid redundant NCC computations
    unique_idx, inv_map = np.unique(ac_idx, return_inverse=True)
    shifts = np.empty(unique_idx.size, dtype=np.float64)
    for j, i in enumerate(unique_idx):
        shift, _ = ncc_tof_shift(ac["wf"][i], ref, dt)
        shifts[j] = shift
    return shifts[inv_map]


def compute_attenuation_per_cycler(ac: dict, ac_idx: np.ndarray,
                                    win_us_start: float, win_us_end: float,
                                    ) -> tuple[np.ndarray, np.ndarray]:
    """Per cycler row, peak-to-peak amplitude and integrated energy of the
    back-wall echo window (specified in µs absolute, sync-edge origin)."""
    fs = ac["fs"]
    gate_start = ac["gate_us_start"]
    # Convert absolute-µs to sample indices within the stored gate
    s0 = max(0, int(round((win_us_start - gate_start) * 1e-6 * fs)))
    s1 = min(ac["wf"].shape[1], int(round((win_us_end - gate_start) * 1e-6 * fs)))
    if s1 <= s0:
        # Window completely outside the gate — fall back to whole gate
        s0, s1 = 0, ac["wf"].shape[1]
    unique_idx, inv_map = np.unique(ac_idx, return_inverse=True)
    p2p = np.empty(unique_idx.size, dtype=np.float64)
    eng = np.empty(unique_idx.size, dtype=np.float64)
    for j, i in enumerate(unique_idx):
        w = ac["wf"][i, s0:s1]
        p2p[j] = float(w.max() - w.min())
        eng[j] = float(np.dot(w, w))
    return p2p[inv_map], eng[inv_map]


# =============================================================================
# Plots
# =============================================================================

def _save(fig: plt.Figure, path: str) -> str:
    fig.savefig(path, bbox_inches="tight")
    plt.close(fig)
    return os.path.abspath(path)


def plot_voltage_current(t_h: np.ndarray, v: np.ndarray, i_ma: np.ndarray,
                          path: str) -> str:
    fig, ax = plt.subplots(figsize=(11, 4.5), dpi=140)
    ax.plot(t_h, v, color="tab:blue", linewidth=1.1, label="Voltage")
    ax.set_xlabel("elapsed time (h)")
    ax.set_ylabel("Voltage (V)", color="tab:blue")
    ax.tick_params(axis="y", labelcolor="tab:blue")
    ax.grid(True, alpha=0.3)
    ax2 = ax.twinx()
    ax2.plot(t_h, i_ma, color="tab:gray", linewidth=1.1, alpha=0.7, label="Current")
    ax2.axhline(0, color="tab:gray", linewidth=0.4, alpha=0.3)
    ax2.set_ylabel("Current (mA)", color="tab:gray")
    ax2.tick_params(axis="y", labelcolor="tab:gray")
    ax.set_title("Diagnostic 1: Voltage and current vs time")
    fig.tight_layout()
    return _save(fig, path)


def plot_temperature(t_h: np.ndarray, temp_c: np.ndarray, path: str) -> str:
    fig, ax = plt.subplots(figsize=(11, 3.5), dpi=140)
    ax.plot(t_h, temp_c, color="tab:orange", linewidth=1.0)
    ax.set_xlabel("elapsed time (h)")
    ax.set_ylabel("Temperature (°C)")
    ax.set_title(f"Diagnostic 2: Temperature vs time  "
                 f"(range {temp_c.max()-temp_c.min():.2f} °C)")
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    return _save(fig, path)


def plot_tof_vs_time(t_h: np.ndarray, tof_ns: np.ndarray, path: str) -> str:
    fig, ax = plt.subplots(figsize=(11, 4.5), dpi=140)
    ax.plot(t_h, tof_ns, color="tab:red", linewidth=1.0)
    ax.axhline(0, color="black", linewidth=0.4, alpha=0.5)
    ax.set_xlabel("elapsed time (h)")
    ax.set_ylabel("ToF shift (ns)  [NCC vs reference scan]")
    ax.set_title("Diagnostic 3: ToF shift vs time")
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    return _save(fig, path)


def plot_tof_vs_soc(soc: np.ndarray, tof_ns: np.ndarray, i_a: np.ndarray,
                    path: str, sensitivity: dict) -> str:
    """Headline plot: ToF vs SoC, charge and discharge plotted separately."""
    fig, ax = plt.subplots(figsize=(10, 6), dpi=150)

    is_charge    = i_a > 1e-6
    is_discharge = i_a < -1e-6

    if is_charge.any():
        ax.scatter(soc[is_charge], tof_ns[is_charge],
                   s=8, color="tab:red", alpha=0.6, label="Charge")
    if is_discharge.any():
        ax.scatter(soc[is_discharge], tof_ns[is_discharge],
                   s=8, color="tab:blue", alpha=0.6, label="Discharge")

    # Overlay the cubic fit used in the noise summary
    if sensitivity.get("ok") and sensitivity.get("cubic_coeffs") is not None:
        soc_fit = np.linspace(sensitivity["soc_charge_min"],
                              sensitivity["soc_charge_max"], 400)
        tof_fit = np.polyval(sensitivity["cubic_coeffs"], soc_fit)
        ax.plot(soc_fit, tof_fit, color="black", linestyle="--", linewidth=1.0,
                label=(f"Cubic fit  |  max d/dSoC = "
                       f"{sensitivity['max_sensitivity_ns_per_pct']:.2f} ns/%SoC"))

    ax.set_xlabel("SoC (%)")
    ax.set_ylabel("ToF shift (ns)  [NCC vs reference scan]")
    ax.set_title("HEADLINE: ToF vs SoC")
    ax.grid(True, alpha=0.3)
    ax.legend(loc="best", fontsize=9)
    fig.tight_layout()
    return _save(fig, path)


def plot_attenuation_vs_soc(soc: np.ndarray, p2p: np.ndarray, energy: np.ndarray,
                             i_a: np.ndarray, path: str) -> str:
    fig, axes = plt.subplots(1, 2, figsize=(13, 5), dpi=140)
    is_charge = i_a > 1e-6
    for ax, y, name in [(axes[0], p2p, "Peak-to-peak amplitude (V)"),
                         (axes[1], energy, "Windowed energy")]:
        if is_charge.any():
            ax.scatter(soc[is_charge], y[is_charge], s=8, color="tab:green", alpha=0.6, label="Charge")
        is_discharge = i_a < -1e-6
        if is_discharge.any():
            ax.scatter(soc[is_discharge], y[is_discharge], s=8, color="tab:purple", alpha=0.6, label="Discharge")
        ax.set_xlabel("SoC (%)")
        ax.set_ylabel(name)
        ax.grid(True, alpha=0.3)
        ax.legend(loc="best", fontsize=9)
    fig.suptitle("Diagnostic 5: Attenuation vs SoC", y=1.02)
    fig.tight_layout()
    return _save(fig, path)


def plot_cycle_drift(scan_ref: np.ndarray, scan_last: np.ndarray,
                      fs: float, gate_us_start: float, path: str) -> str:
    """Compare first and last A-scan to spot coupling drift over the cycle."""
    x_us = gate_us_start + np.arange(scan_ref.size) / fs * 1e6
    fig, ax = plt.subplots(figsize=(11, 4.5), dpi=140)
    ax.plot(x_us, scan_ref,  color="tab:blue",  linewidth=1.0, label="Reference scan (start)")
    ax.plot(x_us, scan_last, color="tab:red",   linewidth=1.0, alpha=0.7, label="Last cycler-paired scan")
    ax.set_xlabel("time after sync (µs)")
    ax.set_ylabel("voltage (V)")
    ax.set_title("Diagnostic 6: First vs last paired A-scan  (coupling-drift check)")
    ax.grid(True, alpha=0.3)
    ax.legend()
    fig.tight_layout()
    return _save(fig, path)


# =============================================================================
# Noise / sensitivity quantification
# =============================================================================

def quantify_sensitivity(soc: np.ndarray, tof_ns: np.ndarray, i_a: np.ndarray
                          ) -> dict:
    """Quantify ToF measurement noise and SoC sensitivity for the charge phase.

    Reports four numbers:
      - sigma_tof_ns: std of residuals after subtracting a smooth baseline
        (rolling median over 50 SoC samples). This is the *measurement* noise,
        NOT the deviation from a line — important for LFP, where the
        ToF-vs-SoC relationship is intrinsically non-linear/non-monotonic.
      - max_sensitivity_ns_per_pct: |slope| at the steepest section of the
        charge curve, from a cubic polynomial fit (gives the best-case
        resolving power).
      - mean_sensitivity_ns_per_pct: |slope| averaged over the charge curve.
      - min_detectable_soc_pct: 2·sigma / max_sensitivity = the smallest SoC
        change resolvable at the best section of the curve.
    Also returns the cubic polynomial coefficients so the headline plot can
    overlay the fit.
    """
    mask = (i_a > 1e-6) & np.isfinite(soc) & np.isfinite(tof_ns)
    if mask.sum() < 20:
        return {"n_points": int(mask.sum()), "ok": False}

    # Sort by SoC for clean derivative computation
    order = np.argsort(soc[mask])
    s = soc[mask][order]
    y = tof_ns[mask][order]

    # Smooth baseline: rolling median over ~50 samples (~0.7 % SoC for C/10)
    # gives a curve following the physics, leaving residuals as pure noise.
    win = min(51, max(5, (len(s) // 50) | 1))   # odd window, max 51
    smoothed = np.empty_like(y)
    half = win // 2
    for k in range(len(y)):
        lo = max(0, k - half)
        hi = min(len(y), k + half + 1)
        smoothed[k] = np.median(y[lo:hi])
    residual = y - smoothed
    sigma = float(np.std(residual))

    # Sensitivity: derivative of a cubic fit over the charge phase
    coeffs = np.polyfit(s, y, 3)                # cubic in SoC
    deriv  = np.polyder(np.poly1d(coeffs))
    soc_grid = np.linspace(s.min(), s.max(), 400)
    d_grid   = np.abs(deriv(soc_grid))
    max_sens  = float(d_grid.max())
    mean_sens = float(d_grid.mean())

    min_resolvable = (2.0 * sigma / max_sens) if max_sens > 1e-12 else float("inf")

    return {
        "ok":                            True,
        "n_points":                      int(mask.sum()),
        "sigma_tof_ns":                  sigma,
        "max_sensitivity_ns_per_pct":    max_sens,
        "mean_sensitivity_ns_per_pct":   mean_sens,
        "min_detectable_soc_pct":        min_resolvable,
        "snr_sensitivity":               float(max_sens / max(sigma, 1e-12)),
        "cubic_coeffs":                  coeffs.tolist(),
        "soc_charge_min":                float(s.min()),
        "soc_charge_max":                float(s.max()),
    }


# =============================================================================
# Main
# =============================================================================

def run(h5_path: str, cycler_xlsx: str, temperature_csv: str,
        q_nominal_ah: float = 0.86,
        soc_initial_pct: float = 0.0,
        ref_snapshot: int = 0,
        echo_window_us: tuple[float, float] | None = None,
        cycler_offset_s: float = 0.0,
        temp_offset_s: float = 0.0,
        out_prefix: str = "") -> dict:
    print(f"=== {Path(h5_path).name} + cycler + temperature ===")
    ac = read_acoustic(h5_path)
    cy = read_cycler_xlsx(cycler_xlsx)
    tp = read_temperature_csv(temperature_csv)
    print(f"  acoustic:    {len(ac['t_s']):,} snapshots  duration={ac['t_s'][-1]/3600:.2f}h  "
          f"gate={ac['gate_us_start']}-{ac['gate_us_end']}us  fs={ac['fs']/1e6:.1f}MHz")
    print(f"  cycler:      {len(cy['t_s']):,} rows  duration={cy['t_s'][-1]/3600:.2f}h  "
          f"V=[{cy['v'].min():.3f},{cy['v'].max():.3f}]V  "
          f"I=[{cy['i'].min()*1000:.1f},{cy['i'].max()*1000:.1f}]mA")
    print(f"  temperature: {len(tp['t_s']):,} rows  duration={tp['t_s'][-1]/3600:.2f}h  "
          f"T=[{tp['temp'].min():.2f},{tp['temp'].max():.2f}]C")
    print()

    # 1) Synchronise
    sync = synchronise(ac, cy, tp, cycler_offset_s, temp_offset_s)

    # 2) Coulomb-counted SoC
    soc = coulomb_soc(cy["t_s"], cy["i"], q_nominal_ah, soc_initial_pct,
                      charge_positive=True)

    # 3a) ToF via NCC
    tof_s   = compute_tof_shift_per_cycler(ac, sync["ac_idx"], ref_snapshot)
    tof_ns  = tof_s * 1e9

    # 3b) Attenuation
    if echo_window_us is None:
        echo_window_us = (ac["gate_us_start"], ac["gate_us_end"])
    p2p, eng = compute_attenuation_per_cycler(ac, sync["ac_idx"],
                                              echo_window_us[0], echo_window_us[1])

    # 4) Plots
    prefix = out_prefix or f"cycle_{Path(h5_path).stem}"
    t_h = sync["t_cycler_h"]
    paths = []
    paths.append(plot_voltage_current(t_h, cy["v"], cy["i"] * 1000.0,
                                       f"{prefix}_diag_voltage_current.png"))
    paths.append(plot_temperature(t_h, sync["temp_on_cy"],
                                   f"{prefix}_diag_temperature.png"))
    paths.append(plot_tof_vs_time(t_h, tof_ns,
                                   f"{prefix}_diag_tof_vs_time.png"))

    # 5) Sensitivity / noise — compute first so we can overlay the fit
    sensitivity = quantify_sensitivity(soc, tof_ns, cy["i"])

    paths.append(plot_tof_vs_soc(soc, tof_ns, cy["i"],
                                  f"{prefix}_headline_tof_vs_soc.png", sensitivity))
    paths.append(plot_attenuation_vs_soc(soc, p2p, eng, cy["i"],
                                          f"{prefix}_attenuation_vs_soc.png"))
    paths.append(plot_cycle_drift(ac["wf"][ref_snapshot],
                                   ac["wf"][sync["ac_idx"][-1]],
                                   ac["fs"], ac["gate_us_start"],
                                   f"{prefix}_cycle_drift.png"))

    # 5) Noise / sensitivity summary
    print("--- noise / sensitivity summary (charge phase only) ---")
    if not sensitivity.get("ok"):
        print(f"  (not enough charge-phase points: {sensitivity.get('n_points', 0)})")
    else:
        sigma     = sensitivity["sigma_tof_ns"]
        max_sens  = sensitivity["max_sensitivity_ns_per_pct"]
        mean_sens = sensitivity["mean_sensitivity_ns_per_pct"]
        mdr       = sensitivity["min_detectable_soc_pct"]
        snr       = sensitivity["snr_sensitivity"]
        print(f"  fit points:                 {sensitivity['n_points']}")
        print(f"  sigma_ToF (smoothed resid): {sigma:.3f} ns")
        print(f"  max |d(ToF)/d(SoC)|:        {max_sens:.3f} ns / %SoC  "
              f"(steepest section of charge curve)")
        print(f"  mean |d(ToF)/d(SoC)|:       {mean_sens:.3f} ns / %SoC  "
              f"(averaged across charge)")
        print(f"  min resolvable dSoC:        {mdr:.4f} %  (= 2 sigma / max_sensitivity)")
        print(f"  sensitivity / noise:        {snr:.1f}  "
              f"({'GOOD (>10)' if snr > 10 else 'MARGINAL (3-10)' if snr > 3 else 'POOR (<3)'})")

    print()
    print("--- saved plots ---")
    for p in paths:
        print(f"  {p}")

    return {
        "soc":          soc,
        "tof_ns":       tof_ns,
        "p2p":          p2p,
        "energy":       eng,
        "sensitivity":  sensitivity,
        "paths":        paths,
    }


def main() -> None:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("h5", help="A-scan HDF5 file")
    p.add_argument("--cycler", required=True, help="Autolab cycler XLSX")
    p.add_argument("--temp",   required=True, help="PicoLog temperature CSV")
    p.add_argument("--q-nominal-ah",  type=float, default=0.86,
                    help="Nominal cell capacity in Ah (default 0.86 — LFP, C/10 at 86 mA)")
    p.add_argument("--soc-initial-pct", type=float, default=0.0,
                    help="SoC at the first cycler timestamp (default 0)")
    p.add_argument("--ref-snapshot", type=int, default=0,
                    help="A-scan snapshot to use as the NCC reference (default 0)")
    p.add_argument("--echo-window-us", nargs=2, type=float, default=None,
                    metavar=("US_START", "US_END"),
                    help="Back-wall echo window in absolute µs (defaults to full gate)")
    p.add_argument("--cycler-offset", type=float, default=0.0,
                    help="Shift cycler time by this many seconds (default 0)")
    p.add_argument("--temp-offset", type=float, default=0.0,
                    help="Shift temperature time by this many seconds (default 0)")
    p.add_argument("--out-prefix", default="",
                    help="Prefix for output PNG filenames")
    args = p.parse_args()

    run(args.h5,
        args.cycler,
        args.temp,
        q_nominal_ah=args.q_nominal_ah,
        soc_initial_pct=args.soc_initial_pct,
        ref_snapshot=args.ref_snapshot,
        echo_window_us=tuple(args.echo_window_us) if args.echo_window_us else None,
        cycler_offset_s=args.cycler_offset,
        temp_offset_s=args.temp_offset,
        out_prefix=args.out_prefix,
    )


if __name__ == "__main__":
    main()

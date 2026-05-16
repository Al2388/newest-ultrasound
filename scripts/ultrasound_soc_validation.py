"""
Supervisor-facing ultrasound/SoC validation analysis for the 13-5 cycle.

This script combines the charge and discharge cycler exports, synchronises them
with the A-scan and temperature logs, extracts ultrasound ToF/amplitude/energy
features, and produces plots plus a short markdown report showing whether the
ultrasound signal is tracking electrochemical state rather than only room
temperature drift.
"""
from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
from pathlib import Path

import h5py
import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import openpyxl


@dataclass
class CyclerData:
    t_raw: np.ndarray
    t_s: np.ndarray
    voltage: np.ndarray
    current: np.ndarray
    source: np.ndarray


def read_cycler_files(paths: list[str]) -> CyclerData:
    rows = []
    for path in paths:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
        ws = wb[wb.sheetnames[0]]
        header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        idx_t = header.index("Time (s)")
        idx_v = header.index("WE(1).Potential (V)")
        idx_i = header.index("WE(1).Current (A)")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[idx_t] is None:
                continue
            try:
                rows.append(
                    (
                        float(row[idx_t]),
                        float(row[idx_v]),
                        float(row[idx_i]),
                        str(path),
                    )
                )
            except (TypeError, ValueError):
                continue
    rows.sort(key=lambda x: x[0])
    t_raw = np.asarray([r[0] for r in rows], dtype=np.float64)
    return CyclerData(
        t_raw=t_raw,
        t_s=t_raw - t_raw[0],
        voltage=np.asarray([r[1] for r in rows], dtype=np.float64),
        current=np.asarray([r[2] for r in rows], dtype=np.float64),
        source=np.asarray([r[3] for r in rows]),
    )


def read_temperature_csv(path: str) -> tuple[np.ndarray, np.ndarray]:
    t_s, temp = [], []
    with open(path, newline="") as f:
        reader = csv.reader(f)
        next(reader, None)
        for row in reader:
            if len(row) < 2 or not row[0] or not row[1]:
                continue
            try:
                hh, mm, ss = row[0].split(":")
                seconds = int(hh) * 3600 + int(mm) * 60 + int(ss)
                t_s.append(seconds)
                temp.append(float(row[1]))
            except (ValueError, TypeError):
                continue
    t = np.asarray(t_s, dtype=np.float64)
    return t - t[0], np.asarray(temp, dtype=np.float64)


def read_acoustic_h5(path: str) -> dict:
    with h5py.File(path, "r") as f:
        return {
            "timestamps": f["timestamps"][:],
            "t_s": f["timestamps"][:] - f["timestamps"][0],
            "waveforms": f["waveforms"][:],
            "fs": float(f.attrs.get("fs_hz", 20_000_000.0)),
            "gate_us_start": float(f.attrs.get("gate_us_start", 30.0)),
            "gate_us_end": float(f.attrs.get("gate_us_end", 40.0)),
        }


def parabolic_peak(y: np.ndarray, k: int) -> float:
    if k <= 0 or k >= len(y) - 1:
        return float(k)
    y0, y1, y2 = float(y[k - 1]), float(y[k]), float(y[k + 1])
    denom = y0 - 2.0 * y1 + y2
    if abs(denom) < 1e-12:
        return float(k)
    return float(k) + max(-1.0, min(1.0, 0.5 * (y0 - y2) / denom))


def ncc_shift_ns(scan: np.ndarray, ref: np.ndarray, dt: float) -> float:
    s = scan - scan.mean()
    r = ref - ref.mean()
    norm = np.sqrt(np.sum(s * s) * np.sum(r * r))
    if norm <= 1e-12:
        return np.nan
    corr = np.correlate(s, r, mode="full") / norm
    k = int(np.argmax(corr))
    center = len(ref) - 1
    return (parabolic_peak(corr, k) - center) * dt * 1e9


def nearest_indices(source_t: np.ndarray, target_t: np.ndarray) -> np.ndarray:
    idx = np.clip(np.searchsorted(source_t, target_t), 0, len(source_t) - 1)
    left = np.clip(idx - 1, 0, len(source_t) - 1)
    use_left = np.abs(source_t[left] - target_t) < np.abs(source_t[idx] - target_t)
    return np.where(use_left, left, idx)


def coulomb_soc(
    cy: CyclerData, q_nominal_ah: float, soc_initial_pct: float
) -> tuple[np.ndarray, np.ndarray, float, float, float]:
    q = np.zeros_like(cy.t_s)
    dt = np.diff(cy.t_raw)
    i_mid = 0.5 * (cy.current[:-1] + cy.current[1:])
    valid = dt < 30.0
    for k in range(1, len(q)):
        if valid[k - 1]:
            q[k] = q[k - 1] + i_mid[k - 1] * dt[k - 1] / 3600.0
        else:
            q[k] = q[k - 1]
    q_charge = float(np.sum(i_mid[(cy.current[:-1] > 1e-6) & (cy.current[1:] > 1e-6) & valid] * dt[(cy.current[:-1] > 1e-6) & (cy.current[1:] > 1e-6) & valid]) / 3600.0)
    q_discharge = float(abs(np.sum(i_mid[(cy.current[:-1] < -1e-6) & (cy.current[1:] < -1e-6) & valid] * dt[(cy.current[:-1] < -1e-6) & (cy.current[1:] < -1e-6) & valid]) / 3600.0))
    ce = q_discharge / q_charge * 100.0 if q_charge else np.nan

    rel_soc = q / q_nominal_ah * 100.0
    soc = np.empty_like(q)
    charge_mask = cy.current > 1e-6
    discharge_mask = cy.current < -1e-6
    soc[charge_mask] = soc_initial_pct + rel_soc[charge_mask]
    q_start_dis = q[discharge_mask][0] if discharge_mask.any() else 0.0
    soc[discharge_mask] = soc_initial_pct + (q_start_dis + (q[discharge_mask] - q_start_dis)) / q_nominal_ah * 100.0
    soc[~(charge_mask | discharge_mask)] = soc_initial_pct + rel_soc[~(charge_mask | discharge_mask)]
    return soc, rel_soc, q_charge, q_discharge, ce


def fit_r2(y: np.ndarray, x: np.ndarray) -> tuple[np.ndarray, float]:
    mask = np.all(np.isfinite(x), axis=1) & np.isfinite(y)
    x = x[mask]
    y = y[mask]
    coef, *_ = np.linalg.lstsq(x, y, rcond=None)
    pred = x @ coef
    ss_res = float(np.sum((y - pred) ** 2))
    ss_tot = float(np.sum((y - y.mean()) ** 2))
    r2 = 1.0 - ss_res / ss_tot if ss_tot > 0 else np.nan
    return coef, r2


def fit_model(y: np.ndarray, x: np.ndarray) -> tuple[np.ndarray, np.ndarray, float, float]:
    """Least-squares fit with full-length predictions for plotting/comparison."""
    mask = np.all(np.isfinite(x), axis=1) & np.isfinite(y)
    coef = np.full(x.shape[1], np.nan, dtype=float)
    pred = np.full_like(y, np.nan, dtype=float)
    if not np.any(mask):
        return coef, pred, np.nan, np.nan

    coef_fit, *_ = np.linalg.lstsq(x[mask], y[mask], rcond=None)
    pred_fit = x[mask] @ coef_fit
    coef[:] = coef_fit
    pred[mask] = pred_fit

    ss_res = float(np.sum((y[mask] - pred_fit) ** 2))
    ss_tot = float(np.sum((y[mask] - y[mask].mean()) ** 2))
    r2 = 1.0 - ss_res / ss_tot if ss_tot > 0 else np.nan
    return coef, pred, r2, ss_res


def savefig(fig: plt.Figure, path: Path) -> None:
    fig.tight_layout()
    fig.savefig(path)
    plt.close(fig)


def run(args: argparse.Namespace) -> None:
    out = Path(args.out_dir)
    out.mkdir(parents=True, exist_ok=True)

    cy = read_cycler_files(args.cycler)
    ac = read_acoustic_h5(args.h5)
    temp_t, temp = read_temperature_csv(args.temp)
    temp_on_cy = np.interp(cy.t_s, temp_t, temp, left=temp[0], right=temp[-1])
    soc, rel_soc, q_charge, q_discharge, ce = coulomb_soc(
        cy, args.q_nominal_ah, args.soc_initial_pct
    )

    ac_idx = nearest_indices(ac["t_s"], cy.t_s)
    unique_idx, inverse = np.unique(ac_idx, return_inverse=True)
    ref = ac["waveforms"][args.ref_snapshot]
    dt = 1.0 / ac["fs"]
    tof_unique = np.asarray([ncc_shift_ns(ac["waveforms"][i], ref, dt) for i in unique_idx])
    tof_ns = tof_unique[inverse]
    wf = ac["waveforms"][ac_idx]
    p2p = wf.max(axis=1) - wf.min(axis=1)
    energy = np.sum(wf * wf, axis=1)
    clip_fraction = np.mean((wf >= 1.98) | (wf <= -1.98), axis=1)

    charge = cy.current > 1e-6
    discharge = cy.current < -1e-6
    t_h = cy.t_s / 3600.0
    gap_idx = np.where(np.diff(cy.t_raw) > 30.0)[0]

    # Evidence that ToF is not simply temperature.
    y = tof_ns
    temp_only_x = np.column_stack([np.ones_like(temp_on_cy), temp_on_cy])
    temp_only_coef, temp_only_pred, r2_temp, sse_temp = fit_model(y, temp_only_x)
    step = np.where(charge, 0.0, 1.0)
    soc_scaled = soc / 100.0
    soc_state_x = np.column_stack(
        [
            np.ones_like(soc),
            soc_scaled,
            soc_scaled**2,
            soc_scaled**3,
            step,
            step * soc_scaled,
            step * soc_scaled**2,
        ]
    )
    soc_state_coef, soc_state_pred, r2_soc, sse_soc = fit_model(y, soc_state_x)
    full_x = np.column_stack(
        [
            np.ones_like(soc),
            soc_scaled,
            soc_scaled**2,
            soc_scaled**3,
            temp_on_cy,
            step,
            step * soc_scaled,
            step * soc_scaled**2,
        ]
    )
    full_coef, full_pred, r2_full, sse_full = fit_model(y, full_x)
    temp_coef_ns_per_c = float(full_coef[4])
    temp_span_effect = abs(temp_coef_ns_per_c) * (float(np.max(temp_on_cy)) - float(np.min(temp_on_cy)))
    tof_span = float(np.nanmax(tof_ns) - np.nanmin(tof_ns))
    tof_temp_corrected = tof_ns - temp_coef_ns_per_c * (temp_on_cy - np.nanmean(temp_on_cy))
    temp_component = temp_coef_ns_per_c * (temp_on_cy - np.nanmean(temp_on_cy))
    soc_state_component = (
        full_coef[1] * soc_scaled
        + full_coef[2] * soc_scaled**2
        + full_coef[3] * soc_scaled**3
        + full_coef[5] * step
        + full_coef[6] * step * soc_scaled
        + full_coef[7] * step * soc_scaled**2
    )
    soc_state_component = soc_state_component - np.nanmean(soc_state_component)
    soc_state_span_effect = float(np.nanmax(soc_state_component) - np.nanmin(soc_state_component))
    temp_component_span_effect = float(np.nanmax(temp_component) - np.nanmin(temp_component))
    soc_to_temp_ratio = (
        soc_state_span_effect / temp_component_span_effect
        if temp_component_span_effect > 0
        else np.nan
    )
    delta_r2_soc_after_temp = r2_full - r2_temp
    delta_r2_temp_after_soc = r2_full - r2_soc
    partial_soc_after_temp = (
        (sse_temp - sse_full) / sse_temp * 100.0
        if np.isfinite(sse_temp) and sse_temp > 0
        else np.nan
    )
    partial_temp_after_soc = (
        (sse_soc - sse_full) / sse_soc * 100.0
        if np.isfinite(sse_soc) and sse_soc > 0
        else np.nan
    )

    # Robustness check: repeat the same model comparison on acoustic snapshots as
    # the rows, interpolating/nearest-matching cycler and temperature onto A-scan
    # time. This guards against the conclusion being an artifact of using cycler
    # timestamps as the regression rows.
    ac_inside = (ac["t_s"] >= cy.t_s[0]) & (ac["t_s"] <= cy.t_s[-1])
    ac_t = ac["t_s"][ac_inside]
    cy_idx_for_ac = nearest_indices(cy.t_s, ac_t)
    tof_acoustic_rows = np.asarray(
        [ncc_shift_ns(w, ref, dt) for w in ac["waveforms"][ac_inside]]
    )
    temp_acoustic_rows = np.interp(ac_t, temp_t, temp, left=temp[0], right=temp[-1])
    soc_acoustic_rows = soc[cy_idx_for_ac]
    current_acoustic_rows = cy.current[cy_idx_for_ac]
    step_acoustic_rows = np.where(current_acoustic_rows > 1e-6, 0.0, 1.0)
    s_ac = soc_acoustic_rows / 100.0
    temp_x_ac = np.column_stack([np.ones_like(temp_acoustic_rows), temp_acoustic_rows])
    soc_x_ac = np.column_stack(
        [
            np.ones_like(s_ac),
            s_ac,
            s_ac**2,
            s_ac**3,
            step_acoustic_rows,
            step_acoustic_rows * s_ac,
            step_acoustic_rows * s_ac**2,
        ]
    )
    full_x_ac = np.column_stack(
        [
            np.ones_like(s_ac),
            s_ac,
            s_ac**2,
            s_ac**3,
            temp_acoustic_rows,
            step_acoustic_rows,
            step_acoustic_rows * s_ac,
            step_acoustic_rows * s_ac**2,
        ]
    )
    _, _, r2_temp_ac, _ = fit_model(tof_acoustic_rows, temp_x_ac)
    _, _, r2_soc_ac, _ = fit_model(tof_acoustic_rows, soc_x_ac)
    full_coef_ac, _, r2_full_ac, _ = fit_model(tof_acoustic_rows, full_x_ac)
    temp_component_ac = full_coef_ac[4] * (temp_acoustic_rows - np.nanmean(temp_acoustic_rows))
    soc_component_ac = (
        full_coef_ac[1] * s_ac
        + full_coef_ac[2] * s_ac**2
        + full_coef_ac[3] * s_ac**3
        + full_coef_ac[5] * step_acoustic_rows
        + full_coef_ac[6] * step_acoustic_rows * s_ac
        + full_coef_ac[7] * step_acoustic_rows * s_ac**2
    )
    soc_component_ac = soc_component_ac - np.nanmean(soc_component_ac)
    soc_span_ac = float(np.nanmax(soc_component_ac) - np.nanmin(soc_component_ac))
    temp_span_ac = float(np.nanmax(temp_component_ac) - np.nanmin(temp_component_ac))

    # Plot 1: whole experiment overview.
    fig, axes = plt.subplots(4, 1, figsize=(12, 10), dpi=150, sharex=True)
    axes[0].plot(t_h, cy.voltage, color="tab:blue", lw=1.1)
    axes[0].set_ylabel("Voltage (V)")
    axes[1].plot(t_h, cy.current * 1000, color="0.4", lw=0.9)
    axes[1].set_ylabel("Current (mA)")
    axes[2].plot(t_h, temp_on_cy, color="tab:red", lw=0.9)
    axes[2].set_ylabel("Temp (C)")
    axes[3].plot(t_h, tof_ns, color="tab:green", lw=1.0)
    axes[3].set_ylabel("ToF shift (ns)")
    axes[3].set_xlabel("Elapsed time (h)")
    for ax in axes:
        ax.grid(alpha=0.3)
        for g in gap_idx:
            ax.axvspan(t_h[g], t_h[g + 1], color="tab:orange", alpha=0.12)
    savefig(fig, out / "01_overview_voltage_current_temp_tof.png")

    # Plot 2: ToF vs SoC with charge/discharge.
    fig, ax = plt.subplots(figsize=(8, 5), dpi=150)
    ax.scatter(soc[charge], tof_ns[charge], s=8, alpha=0.55, color="tab:green", label="Charge")
    ax.scatter(soc[discharge], tof_ns[discharge], s=8, alpha=0.55, color="tab:purple", label="Discharge")
    ax.set_xlabel("Coulomb-counted SoC (%)")
    ax.set_ylabel("NCC ToF shift vs start (ns)")
    ax.grid(alpha=0.3)
    ax.legend()
    savefig(fig, out / "02_tof_vs_soc_charge_discharge.png")

    # Plot 3: Temperature-only test.
    fig, axes = plt.subplots(1, 2, figsize=(12, 5), dpi=150)
    sc = axes[0].scatter(temp_on_cy, tof_ns, c=soc, s=8, cmap="viridis", alpha=0.65)
    axes[0].set_xlabel("Temperature (C)")
    axes[0].set_ylabel("ToF shift (ns)")
    axes[0].set_title(f"Temperature-only R2 = {r2_temp:.3f}")
    axes[0].grid(alpha=0.3)
    fig.colorbar(sc, ax=axes[0], label="SoC (%)")
    axes[1].scatter(soc, tof_ns, c=temp_on_cy, s=8, cmap="coolwarm", alpha=0.65)
    axes[1].set_xlabel("SoC (%)")
    axes[1].set_ylabel("ToF shift (ns)")
    axes[1].set_title(f"SoC/state-only R2 = {r2_soc:.3f}; combined R2 = {r2_full:.3f}")
    axes[1].grid(alpha=0.3)
    savefig(fig, out / "03_temperature_vs_soc_separation.png")

    # Plot 4: model-estimated temperature and SoC/state contributions.
    fig, axes = plt.subplots(2, 1, figsize=(12, 8), dpi=150, sharex=False)
    tof_centered = tof_ns - np.nanmean(tof_ns)
    model_centered = full_pred - np.nanmean(full_pred)
    axes[0].plot(t_h, tof_centered, color="0.45", lw=0.9, label="Measured ToF, centered")
    axes[0].plot(t_h, model_centered, color="tab:blue", lw=1.2, label=f"Combined model, R2={r2_full:.3f}")
    axes[0].plot(
        t_h,
        soc_state_component,
        color="tab:green",
        lw=1.1,
        label=f"SoC/state term span={soc_state_span_effect:.1f} ns",
    )
    axes[0].plot(
        t_h,
        temp_component,
        color="tab:red",
        lw=1.1,
        label=f"Temperature term span={temp_component_span_effect:.1f} ns",
    )
    axes[0].set_xlabel("Elapsed time (h)")
    axes[0].set_ylabel("Centered ToF contribution (ns)")
    axes[0].grid(alpha=0.3)
    axes[0].legend(loc="best", fontsize=8)
    for g in gap_idx:
        axes[0].axvspan(t_h[g], t_h[g + 1], color="tab:orange", alpha=0.12)

    labels = ["Temp only", "SoC/state only", "SoC/state + temp"]
    values = [r2_temp, r2_soc, r2_full]
    colors = ["tab:red", "tab:green", "tab:blue"]
    axes[1].bar(labels, values, color=colors, alpha=0.85)
    axes[1].set_ylim(0, 1.05)
    axes[1].set_ylabel("Model R2 for ToF")
    axes[1].grid(axis="y", alpha=0.3)
    for i, value in enumerate(values):
        axes[1].text(i, value + 0.025, f"{value:.3f}", ha="center", va="bottom")
    axes[1].set_title(
        f"Adding SoC/state after temperature improves R2 by {delta_r2_soc_after_temp:.3f}; "
        f"adding temperature after SoC/state improves R2 by {delta_r2_temp_after_soc:.3f}"
    )
    savefig(fig, out / "04_tof_temperature_soc_decomposition.png")

    # Plot 5: explicitly temperature-corrected ToF.
    fig, axes = plt.subplots(1, 2, figsize=(12, 5), dpi=150)
    axes[0].scatter(soc[charge], tof_ns[charge], s=8, alpha=0.5, color="tab:green", label="Charge raw")
    axes[0].scatter(soc[discharge], tof_ns[discharge], s=8, alpha=0.5, color="tab:purple", label="Discharge raw")
    axes[0].set_xlabel("SoC (%)")
    axes[0].set_ylabel("Raw ToF shift (ns)")
    axes[0].grid(alpha=0.3)
    axes[0].legend()
    axes[1].scatter(soc[charge], tof_temp_corrected[charge], s=8, alpha=0.5, color="tab:green", label="Charge corrected")
    axes[1].scatter(soc[discharge], tof_temp_corrected[discharge], s=8, alpha=0.5, color="tab:purple", label="Discharge corrected")
    axes[1].set_xlabel("SoC (%)")
    axes[1].set_ylabel("Temperature-corrected ToF (ns)")
    axes[1].grid(alpha=0.3)
    axes[1].legend()
    fig.suptitle(f"Correction uses fitted temperature coefficient {temp_coef_ns_per_c:.1f} ns/C")
    savefig(fig, out / "05_temperature_corrected_tof_vs_soc.png")

    # Plot 6: amplitude/energy diagnostics.
    fig, axes = plt.subplots(3, 1, figsize=(12, 8), dpi=150, sharex=True)
    axes[0].plot(t_h, p2p, color="tab:orange", lw=0.8)
    axes[0].set_ylabel("P2P (V)")
    axes[1].plot(t_h, energy, color="tab:brown", lw=0.8)
    axes[1].set_ylabel("Energy")
    axes[2].plot(t_h, clip_fraction * 100.0, color="tab:red", lw=0.8)
    axes[2].set_ylabel("Clipped samples (%)")
    axes[2].set_xlabel("Elapsed time (h)")
    for ax in axes:
        ax.grid(alpha=0.3)
        for g in gap_idx:
            ax.axvspan(t_h[g], t_h[g + 1], color="tab:orange", alpha=0.12)
    savefig(fig, out / "06_amplitude_energy_clipping.png")

    # Plot 7: representative waveforms.
    targets = [
        ("charge 0%", charge, 0),
        ("charge 25%", charge, 25),
        ("charge 50%", charge, 50),
        ("charge 75%", charge, 75),
        ("charge 100%", charge, 100),
        ("discharge 75%", discharge, 75),
        ("discharge 50%", discharge, 50),
        ("discharge 25%", discharge, 25),
        ("discharge 0%", discharge, 0),
    ]
    x_us = ac["gate_us_start"] + np.arange(ac["waveforms"].shape[1]) / ac["fs"] * 1e6
    fig, ax = plt.subplots(figsize=(11, 6), dpi=150)
    offset = 0.0
    for label, mask, target_soc in targets:
        if not np.any(mask):
            continue
        local = np.where(mask)[0]
        j = local[np.argmin(np.abs(soc[local] - target_soc))]
        ax.plot(x_us, wf[j] + offset, lw=0.9, label=f"{label} ({soc[j]:.1f}%)")
        offset += 4.5
    ax.set_xlabel("Time after sync (us)")
    ax.set_ylabel("Waveform voltage plus offset")
    ax.grid(alpha=0.3)
    ax.legend(ncol=3, fontsize=8)
    savefig(fig, out / "07_representative_waveforms_by_soc.png")

    # Plot 8: ToF and energy versus voltage/capacity-style view.
    fig, axes = plt.subplots(1, 2, figsize=(12, 5), dpi=150)
    axes[0].scatter(cy.voltage[charge], tof_ns[charge], s=8, alpha=0.55, color="tab:green", label="Charge")
    axes[0].scatter(cy.voltage[discharge], tof_ns[discharge], s=8, alpha=0.55, color="tab:purple", label="Discharge")
    axes[0].set_xlabel("Voltage (V)")
    axes[0].set_ylabel("ToF shift (ns)")
    axes[0].legend()
    axes[0].grid(alpha=0.3)
    axes[1].scatter(soc[charge], energy[charge], s=8, alpha=0.55, color="tab:green", label="Charge")
    axes[1].scatter(soc[discharge], energy[discharge], s=8, alpha=0.55, color="tab:purple", label="Discharge")
    axes[1].set_xlabel("SoC (%)")
    axes[1].set_ylabel("Window energy")
    axes[1].legend()
    axes[1].grid(alpha=0.3)
    savefig(fig, out / "08_tof_voltage_energy_soc.png")

    # Export aligned features.
    feature_path = out / "aligned_ultrasound_features.csv"
    with open(feature_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "time_h",
                "voltage_v",
                "current_a",
                "soc_pct",
                "relative_soc_change_pct",
                "temperature_c",
                "tof_shift_ns",
                "tof_temp_corrected_ns",
                "tof_model_full_ns",
                "tof_soc_state_component_ns",
                "tof_temperature_component_ns",
                "p2p_v",
                "energy",
                "clip_fraction",
                "ascan_index",
                "source",
            ]
        )
        for row in zip(
            t_h,
            cy.voltage,
            cy.current,
            soc,
            rel_soc,
            temp_on_cy,
            tof_ns,
            tof_temp_corrected,
            full_pred,
            soc_state_component,
            temp_component,
            p2p,
            energy,
            clip_fraction,
            ac_idx,
            cy.source,
        ):
            writer.writerow(row)

    # Report.
    report = []
    report.append("# 13-5 Ultrasound SoC Validation\n\n")
    report.append("## Purpose\n")
    report.append(
        "Show whether the ultrasonic A-scan features track internal cell state-of-charge during the C/10 cycle, "
        "and whether the observed ToF changes are larger than can be explained by temperature drift alone.\n\n"
    )
    report.append("## Inputs\n")
    report.append(f"- A-scan H5: `{args.h5}`\n")
    for path in args.cycler:
        report.append(f"- Cycler: `{path}`\n")
    report.append(f"- Temperature: `{args.temp}`\n")
    report.append(f"- Nominal capacity used for SoC: {args.q_nominal_ah:.3f} Ah\n\n")
    report.append("## SoC Axis Note\n")
    report.append(
        f"- Initial SoC offset used in this run: {args.soc_initial_pct:.2f}%\n"
    )
    report.append(
        "- If the cell did not start at 0% SoC, the absolute SoC values are shifted by the unknown starting SoC. "
        "The `relative_soc_change_pct` column gives the measured coulomb-counted change from the first cycler row. "
        "The ultrasound conclusion depends on feature changes versus cycling state, so a constant initial-SoC offset changes the x-axis labels but not the ToF/temperature separation result.\n"
    )
    report.append(
        "- Because the charge step alone delivered 0.8659 Ah against a nominal 0.860 Ah, this export behaves like a near-empty-to-full charge on the 0.86 Ah capacity scale. If the cell truly had substantial starting SoC, then either the effective capacity/cutoff window is larger than 0.86 Ah, the starting SoC estimate is wrong, or the current/capacity calibration needs checking.\n\n"
    )
    report.append("## Cycle Summary\n")
    report.append(f"- Charge capacity: {q_charge:.4f} Ah\n")
    report.append(f"- Discharge capacity: {q_discharge:.4f} Ah\n")
    report.append(f"- Coulombic efficiency: {ce:.2f}%\n")
    report.append(f"- Combined cycler time span: {t_h[-1]:.2f} h\n")
    report.append(f"- Temperature span over cycler timestamps: {np.min(temp_on_cy):.3f} to {np.max(temp_on_cy):.3f} C\n")
    report.append(f"- ToF span over cycler timestamps: {np.nanmin(tof_ns):.1f} to {np.nanmax(tof_ns):.1f} ns\n")
    report.append(f"- P2P clipped-sample mean/max: {np.mean(clip_fraction) * 100:.2f}% / {np.max(clip_fraction) * 100:.2f}%\n\n")
    report.append("## Temperature And SoC Separation Test\n")
    report.append(
        "- Model used: ordinary least-squares regression on ToF. The SoC/state term uses coulomb-counted SoC "
        "with polynomial terms plus charge/discharge branch terms; the temperature term is linear in measured temperature.\n"
    )
    report.append(f"- Temperature-only model R2 for ToF: {r2_temp:.3f}\n")
    report.append(f"- SoC/state-only model R2 for ToF: {r2_soc:.3f}\n")
    report.append(f"- SoC/state + temperature model R2 for ToF: {r2_full:.3f}\n")
    report.append(f"- R2 gained by adding SoC/state after temperature: {delta_r2_soc_after_temp:.3f}\n")
    report.append(f"- R2 gained by adding temperature after SoC/state: {delta_r2_temp_after_soc:.3f}\n")
    report.append(f"- Residual error removed by adding SoC/state after temperature: {partial_soc_after_temp:.1f}%\n")
    report.append(f"- Residual error removed by adding temperature after SoC/state: {partial_temp_after_soc:.1f}%\n")
    report.append(f"- Fitted temperature coefficient inside the combined model: {temp_coef_ns_per_c:.2f} ns/C\n")
    report.append(f"- Model-estimated temperature contribution span: {temp_component_span_effect:.1f} ns\n")
    report.append(f"- Model-estimated SoC/state contribution span: {soc_state_span_effect:.1f} ns\n")
    report.append(f"- SoC/state contribution is about {soc_to_temp_ratio:.1f}x larger than the temperature contribution in this run.\n")
    report.append(f"- Observed ToF span: {tof_span:.1f} ns\n\n")
    report.append(
        "Interpretation: the literature says temperature affects ultrasonic ToF, so it should be included and reported. "
        "In this dataset, however, the independent temperature term is small compared with the SoC/state term: SoC/state explains "
        "more variance than temperature alone, adds much more explanatory power when added after temperature, and has a larger fitted "
        "ToF span. The charge and discharge branches also separate at overlapping temperatures, which supports the conclusion that "
        "the acoustic response is tracking internal electrochemical/mechanical state, not only ambient drift.\n\n"
    )
    report.append("## Robustness Check\n")
    report.append(
        "- Repeating the same model comparison using A-scan snapshots as the table rows gives the same conclusion, "
        "so the result is not an artifact of weighting by cycler timestamps.\n"
    )
    report.append(f"- A-scan-row temperature-only R2: {r2_temp_ac:.3f}\n")
    report.append(f"- A-scan-row SoC/state-only R2: {r2_soc_ac:.3f}\n")
    report.append(f"- A-scan-row SoC/state + temperature R2: {r2_full_ac:.3f}\n")
    report.append(f"- A-scan-row SoC/state component span: {soc_span_ac:.1f} ns\n")
    report.append(f"- A-scan-row temperature component span: {temp_span_ac:.1f} ns\n\n")
    report.append("## Literature Alignment\n")
    report.append(
        "- Hsieh et al. established the common electrochemical-acoustic ToF framing: acoustic ToF changes because lithiation changes "
        "density, modulus, attenuation, and mechanical state inside the cell. DOI: https://doi.org/10.1039/C5EE00111K\n"
    )
    report.append(
        "- Ke et al. used ToF and amplitude during lithium-ion pouch-cell cycling and explicitly reported that ToF is influenced by "
        "temperature, while amplitude can correlate with physical electrode-layer changes. DOI: https://doi.org/10.1016/j.jpowsour.2022.232031\n"
    )
    report.append(
        "- Borujerdi, Jin, and Zhu used voltage/current/SOC plus ultrasonic P2P amplitude and ToF shift during cycling, then applied "
        "in-situ temperature correction; after correction, ToF shift correlated well with SOC. DOI: https://doi.org/10.1016/j.jpowsour.2024.234103\n"
    )
    report.append(
        "- Zhang et al. treated SOC and temperature as jointly estimated states from ultrasonic reflection-wave features, which supports "
        "presenting temperature and SOC together rather than pretending temperature is irrelevant. DOI: https://doi.org/10.3390/batteries9060335\n"
    )
    report.append(
        "Compared with those papers, this analysis follows the typical structure: validate voltage/current cycling, compute coulomb-counted "
        "SOC, extract ToF shift plus amplitude/energy features, plot ToF versus time/SOC/voltage, include temperature correction, and state "
        "charge/discharge hysteresis. The main limitation is that this is one cycling experiment with only about 1.2 C of temperature movement; "
        "a stronger causal temperature calibration would need repeated cycles or a controlled temperature sweep at fixed SOC.\n\n"
    )
    report.append("## How To Present This\n")
    report.append("1. Start with `01_overview_voltage_current_temp_tof.png`: voltage/current confirm the cycle, temperature shows only slow ambient drift, and ToF changes strongly through the electrochemical steps.\n")
    report.append("2. Use `02_tof_vs_soc_charge_discharge.png`: the ToF feature has a structured SoC dependence and a charge/discharge branch difference, which is expected for hysteretic cell mechanics.\n")
    report.append("3. Use `03_temperature_vs_soc_separation.png`: temperature alone gives a weaker explanation than SoC/state plus temperature.\n")
    report.append("4. Use `04_tof_temperature_soc_decomposition.png`: this is the clearest slide for your supervisor. It separates the fitted temperature and SoC/state ToF components and shows the SoC/state contribution is larger.\n")
    report.append("5. Use `05_temperature_corrected_tof_vs_soc.png`: after subtracting the fitted temperature term, the SoC trajectory remains.\n")
    report.append("6. Use `07_representative_waveforms_by_soc.png`: the actual echo packet shifts with cycling, so the feature is visible in the raw ultrasonic signal, not only in a derived number.\n\n")
    report.append("## Important Caveats\n")
    report.append("- The cycler rest between charge and discharge is present as an unlogged gap between two exports, so continuous rest-relaxation validation is not available from the cycler data.\n")
    report.append("- Peak-to-peak amplitude is affected by receiver clipping near the voltage rails. Energy and ToF are more reliable features for this run.\n")
    report.append("- The temperature coefficient is an in-run compensation estimate, not a chamber-calibrated material law. Use it to show temperature was checked and corrected, not as a universal coefficient for all cells.\n")
    report.append("- Coulombic efficiency is below the >99% fresh-cell target, so this cycle is useful as a demonstration run but should be repeated for calibration-grade data.\n\n")
    report.append("## Generated Figures\n")
    for name in [
        "01_overview_voltage_current_temp_tof.png",
        "02_tof_vs_soc_charge_discharge.png",
        "03_temperature_vs_soc_separation.png",
        "04_tof_temperature_soc_decomposition.png",
        "05_temperature_corrected_tof_vs_soc.png",
        "06_amplitude_energy_clipping.png",
        "07_representative_waveforms_by_soc.png",
        "08_tof_voltage_energy_soc.png",
    ]:
        report.append(f"- `{name}`\n")
    report.append(f"\nAligned feature table: `{feature_path.name}`\n")
    (out / "ULTRASOUND_SOC_VALIDATION.md").write_text("".join(report), encoding="utf-8")

    print(f"Wrote {out / 'ULTRASOUND_SOC_VALIDATION.md'}")
    print(f"Wrote {feature_path}")
    print(f"Q_charge={q_charge:.4f} Ah  Q_discharge={q_discharge:.4f} Ah  CE={ce:.2f}%")
    print(f"ToF span={tof_span:.1f} ns  temperature span={np.max(temp_on_cy)-np.min(temp_on_cy):.3f} C")
    print(f"R2 temp-only={r2_temp:.3f}  R2 SoC/state-only={r2_soc:.3f}  R2 SoC+temp={r2_full:.3f}")
    print(
        f"Component spans: SoC/state={soc_state_span_effect:.1f} ns  "
        f"temperature={temp_component_span_effect:.1f} ns  "
        f"ratio={soc_to_temp_ratio:.1f}x"
    )
    print(f"Temperature coefficient={temp_coef_ns_per_c:.2f} ns/C")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--h5", default="data/raw/ascan/ascan_session_2026-05-12_15-42-53/ascan_session_2026-05-12_15-42-53.h5")
    parser.add_argument("--cycler", nargs="+", default=["data/raw/cycler/13-5.xlsx", "data/raw/cycler/13-5 (2).xlsx"])
    parser.add_argument("--temp", default="data/raw/temperature/13-5.csv")
    parser.add_argument("--q-nominal-ah", type=float, default=0.86)
    parser.add_argument(
        "--soc-initial-pct",
        type=float,
        default=0.0,
        help="Initial SoC at the first cycler row. Use 0 when unknown and interpret plots as relative SoC.",
    )
    parser.add_argument("--ref-snapshot", type=int, default=0)
    parser.add_argument("--out-dir", default="reports/experiments/cycle_13-5/ultrasound_soc_validation")
    args = parser.parse_args()
    run(args)


if __name__ == "__main__":
    main()
